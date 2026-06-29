"""Excel 数据加载与 VF 占比匹配"""
import openpyxl
import pandas as pd
import numpy as np
import os

SENTIMENT_MAP = {
    '一刀切': 1, '部分限制': 2, '科学管控': 3,
    '一刀切;暂停': 1, '暂停': 1,
}
SENTIMENT_VALUES = {1, 2, 3}
REGION_MAP = {
    "北京": "华北", "天津": "华北", "河北": "华北", "山西": "华北", "内蒙古": "华北",
    "上海": "华东", "江苏": "华东", "浙江": "华东", "安徽": "华东",
    "福建": "华东", "江西": "华东", "山东": "华东",
    "广东": "华南", "广西": "华南", "海南": "华南",
    "河南": "华中", "湖北": "华中", "湖南": "华中",
    "重庆": "西南", "四川": "西南", "贵州": "西南", "云南": "西南",
    "陕西": "西北", "甘肃": "西北", "宁夏": "西北", "新疆": "西北",
    "辽宁": "东北", "吉林": "东北", "黑龙江": "东北",
    "西藏": "西南",
}

MONTH_COLS = [
    '24AP1','24AP2','24AP3','24AP4','24AP5','24AP6',
    '24AP7','24AP8','24AP9','24AP10','24AP11','24AP12',
    '25AP1','25AP2','25AP3','25AP4','25AP5','25AP6',
    '25AP7','25AP8','25AP9','25AP10','25AP11','25AP12',
    '26AP1','26AP2','26AP3','26AP4','26AP5','26AP6',
    '26AP7','26AP8','26AP9','26AP10','26AP11','26AP12',
]


def load_main_data(xlsx_path):
    """读取主数据 Excel（IV-限控tracking + IV-THHtracking 两张分表）
    返回：(sentiment_df, sales_df, hospital_info)
    """
    xls = pd.ExcelFile(xlsx_path)
    assert 'IV-限控tracking' in xls.sheet_names, '缺少 IV-限控tracking 分表'
    assert 'IV-THHtracking' in xls.sheet_names, '缺少 IV-THHtracking 分表'

    sent_raw = pd.read_excel(xlsx_path, sheet_name='IV-限控tracking')
    sales_raw = pd.read_excel(xlsx_path, sheet_name='IV-THHtracking')

    # 提取医院元信息（前2列：code/name，第3列如有则为province）
    id_col, name_col = sent_raw.columns[0], sent_raw.columns[1]
    info_cols = [id_col, name_col]
    has_province = len(sent_raw.columns) >= 3 and 'province' in str(sent_raw.columns[2]).lower()
    if has_province:
        info_cols.append(sent_raw.columns[2])
    hospital_info = sent_raw[info_cols].copy()
    hospital_info.columns = ['code', 'name'] + (['province'] if has_province else [])
    hospital_info = hospital_info.dropna(subset=['code'])
    hospital_info['code'] = hospital_info['code'].astype(str).str.strip()
    hospital_info['name'] = hospital_info['name'].astype(str).str.strip()
    if has_province:
        hospital_info['province'] = hospital_info['province'].astype(str).str.strip()
        hospital_info['region'] = hospital_info['province'].map(REGION_MAP).fillna('其他')
    else:
        hospital_info['province'] = ''
        hospital_info['region'] = ''
    hospital_info = hospital_info[hospital_info['code'] != '']

    # 只保留实际存在的月份列
    available_months = [m for m in MONTH_COLS if m in sent_raw.columns and m in sales_raw.columns]

    # 舆情数据
    sent_cols = [id_col, name_col] + available_months
    sent_df = sent_raw[sent_cols].copy()
    sent_df.columns = ['code', 'name'] + available_months

    # 销量数据
    sales_cols = [id_col, name_col] + available_months
    sales_df = sales_raw[sales_cols].copy()
    sales_df.columns = ['code', 'name'] + available_months

    return sent_df, sales_df, hospital_info, available_months


def normalize_sentiment(val):
    """将舆情值统一映射为 1/2/3，0/null 返回 NaN"""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return np.nan
    if isinstance(val, (int, float)):
        v = int(val)
        return v if v in SENTIMENT_VALUES else np.nan
    s = str(val).strip()
    if s == '0' or s == '':
        return np.nan
    if s in SENTIMENT_MAP:
        return SENTIMENT_MAP[s]
    try:
        v = int(float(s))
        return v if v in SENTIMENT_VALUES else np.nan
    except:
        return np.nan


def normalize_sales(val):
    """销量：保留数值，非数值/None → NaN"""
    if val is None:
        return np.nan
    try:
        v = float(val)
        return v if not np.isnan(v) else np.nan
    except:
        return np.nan


def pivot_to_long(sent_df, sales_df, months, hospital_info=None):
    """宽表转长表：每行 = 一家HP 一个月份"""
    # 构建 code→province/region 元信息表
    meta = {}
    if hospital_info is not None:
        for _, row in hospital_info.iterrows():
            code = str(row['code']).strip()
            meta[code] = {
                'province': row.get('province', ''),
                'region': row.get('region', ''),
            }

    records = []
    for _, row in sent_df.iterrows():
        code = str(row['code']).strip()
        name = str(row.get('name', '')).strip()
        p = meta.get(code, {}).get('province', '')
        r = meta.get(code, {}).get('region', '')
        for m in months:
            s = normalize_sentiment(row.get(m))
            records.append({'code': code, 'name': name, 'province': p, 'region': r,
                            'month': m, 'sentiment': s})
    sent_long = pd.DataFrame(records)

    records = []
    for _, row in sales_df.iterrows():
        code = str(row['code']).strip()
        for m in months:
            v = normalize_sales(row.get(m))
            records.append({'code': code, 'month': m, 'sales': v})
    sales_long = pd.DataFrame(records)

    merged = sent_long.merge(sales_long, on=['code', 'month'], how='outer')
    return merged


def load_vf_data(vf_path, main_path=None):
    """读取 VF 占比表，返回 {HP名称: 占比值}
    优先从主数据文件的 'VF占比' sheet 读取；
    若主文件无该 sheet，则从独立的 vf_path 读取。
    """
    # 优先：主数据文件内嵌 VF占比 sheet
    if main_path and os.path.exists(main_path):
        wb_main = openpyxl.load_workbook(main_path, data_only=True)
        if 'VF占比' in wb_main.sheetnames:
            ws = wb_main['VF占比']
            hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            try:
                name_i = hdr.index('ACCOUNT_NAME') + 1
                vf_i = hdr.index('VF在分子式中的占比') + 1
            except ValueError:
                pass
            else:
                vfmap = {}
                for r in range(2, ws.max_row + 1):
                    nm = ws.cell(r, name_i).value
                    vf = ws.cell(r, vf_i).value
                    if nm:
                        vfmap[str(nm).strip()] = vf if vf is not None else None
                return vfmap

    # 备选：独立 VF 文件
    if vf_path is None or not os.path.exists(vf_path):
        return {}
    wb = openpyxl.load_workbook(vf_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        name_i = hdr.index('ACCOUNT_NAME') + 1
        vf_i = hdr.index('VF在分子式中的占比') + 1
    except ValueError:
        return {}
    vfmap = {}
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, name_i).value
        vf = ws.cell(r, vf_i).value
        if nm:
            vfmap[str(nm).strip()] = vf if vf is not None else None
    return vfmap


def gen_template_excel(output_path):
    """生成标准模板 Excel"""
    months = MONTH_COLS[:36]  # 24AP1 ~ 26AP12
    wb = openpyxl.Workbook()

    # Sheet 1: 舆情追踪
    ws1 = wb.active
    ws1.title = 'IV-限控tracking'
    headers = ['HP Code', 'HP Name', 'Province'] + months
    for c, h in enumerate(headers, 1):
        ws1.cell(1, c, h)
    ws1.cell(2, 1, 'HP00001')
    ws1.cell(2, 2, '示例HP名称')
    ws1.cell(2, 3, '广东')
    for c in range(4, 4 + len(months)):
        ws1.cell(2, c, 3)

    # Sheet 2: 销量追踪
    ws2 = wb.create_sheet('IV-THHtracking')
    for c, h in enumerate(headers, 1):
        ws2.cell(1, c, h)
    ws2.cell(2, 1, 'HP00001')
    ws2.cell(2, 2, '示例HP名称')
    ws2.cell(2, 3, '广东')
    for c in range(4, 4 + len(months)):
        ws2.cell(2, c, 100000)

    # Sheet 3: VF 占比（可选）
    ws3 = wb.create_sheet('VF占比')
    vf_headers = ['ACCOUNT_NAME', 'VF在分子式中的占比']
    for c, h in enumerate(vf_headers, 1):
        ws3.cell(1, c, h)
    ws3.cell(2, 1, '示例HP名称')
    ws3.cell(2, 2, 0.5)

    wb.save(output_path)
    return output_path
